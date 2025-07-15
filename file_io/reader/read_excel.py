import os
import numpy as np

from typing import List, Union
from openpyxl import load_workbook


def read_excel(path: str, work_table: Union[int, List[int], str, List[str]]) -> List[np.ndarray]:
    """
    Reads specified sheets from an Excel file and returns the data as a list of NumPy arrays.

    Inputs:
    --------
        path (str): The file path to the Excel workbook.
        work_table (Union[int, List[int], str, List[str]]): Specifies the sheets to read.
            - int or List[int]: Index/indices of sheets to read.
            - str or List[str]: Name(s) of sheets to read.

    Returns:
    --------
        List[np.ndarray]: 
            A list of NumPy arrays, each representing the data from a sheet.
    """

    # Ensure the Excel file exists
    assert os.path.exists(path), f"File not found at path: {path}"

    # Normalize `work_table` to always be a list
    if isinstance(work_table, (int, str)):
        work_table = [work_table]

    wb_data = []  # List to store data from each sheet
    workbook = load_workbook(path)  # Load the workbook

    # Loop through each sheet in the workbook
    for index, sheetname in enumerate(workbook.sheetnames):
        # Check if the current sheet is specified by name or index in `work_table`
        if index in work_table or sheetname in work_table:
            sheet = workbook[sheetname]

            # Read sheet data row by row
            data = []
            for row in sheet.iter_rows(values_only=True):  # Only values, not cell objects
                data.append(row)

            # Convert sheet data to a NumPy array
            sheet_data = np.array(data)
            wb_data.append(sheet_data)

    if not wb_data:
        raise ValueError(
            f"No sheets matched the specified criteria in `work_table`: {work_table}"
        )

    return wb_data
