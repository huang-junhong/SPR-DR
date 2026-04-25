import os
from typing import Any, List
from openpyxl import Workbook, load_workbook


def write_excel(path: str, sheet_name: str, data: List[List[Any]]) -> bool:
    """
    将数据写入 Excel 文件中的指定 sheet。
    
    Parameters
    --------
    path (str): Excel 文件的完整路径，如果文件不存在则创建；如果存在，则在原文件上写入。
    sheet_name (str): 目标 sheet 的名称，如果该 sheet 已存在，则会自动添加后缀来保证名称唯一。
    data (list[list[any]]): 要写入的数据，每个子列表代表 Excel 中的一行数据。
    """
    try:
        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            wb = Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
        
        original_sheet_name = sheet_name
        count = 1
        
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_sheet_name}_{count}"
            count += 1

        ws = wb.create_sheet(title=sheet_name)

        for row in data:
            ws.append(row)

        wb.save(path)
        return True
    except Exception as e:
        print(str(e))
        return False